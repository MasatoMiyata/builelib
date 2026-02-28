import openpyxl as px
import csv
import glob

def write_list_2d(sheet, l_2d, start_row, start_col):
    """
    リストをExcelファイルに書き込むための関数
    https://note.nkmk.me/python-openpyxl-usage/
    """
    for y, row in enumerate(l_2d):
        for x, cell in enumerate(row):
            sheet.cell(row=start_row + y,
                        column=start_col + x,
                        value=l_2d[y][x])

def csv2excel(csv_directory, output_filename):
    """
    csv_directory  : Excel化するCSVファイルが保存されたディレクトリを指定
    output_filename: 出力するExcelファイルの名称（拡張子なし）
    """

    # CSVファイルのリスト
    csvfile_list = glob.glob(csv_directory + "/*.csv")

    if len(csvfile_list) == 0:
        raise Exception("CSVファイルが見つかりません。")

    # テンプレートファイル（Excelファイル）を読み込む
    excel_workbook = px.load_workbook('./src/builelib/subtools/inputSheet_template.xlsm',read_only=False, keep_vba=True)

    # デフォルト値
    excel_workbook['0) 基本情報'].cell(row=9,  column=3, value="test") # 建物名称
    excel_workbook['0) 基本情報'].cell(row=12, column=3, value="6")    # 地域区分
    excel_workbook['0) 基本情報'].cell(row=17, column=3, value="100")  # 延べ面積

    for csvfile_name in csvfile_list:
        
        # CSVファイルを読み込む
        with open(csvfile_name , newline='', encoding='cp932', errors='ignore') as data:
            csvdata = list(csv.reader(data, delimiter=','))

        if "様式0" in csvfile_name:

            excel_workbook['0) 基本情報'].cell(row=9,  column=3, value=csvdata[8][2]) # 建物名称
            excel_workbook['0) 基本情報'].cell(row=10, column=4, value=csvdata[9][3]) # 都道府県

            if len(csvdata[9]) >= 6:
                excel_workbook['0) 基本情報'].cell(row=10, column=6, value=csvdata[9][5]) # 市区町村
            excel_workbook['0) 基本情報'].cell(row=11, column=3, value=csvdata[10][2])  # 住所

            excel_workbook['0) 基本情報'].cell(row=12, column=3, value=csvdata[11][2])  # 地域区分
            excel_workbook['0) 基本情報'].cell(row=13, column=3, value=csvdata[12][2])  # 構造

            excel_workbook['0) 基本情報'].cell(row=14, column=4, value=csvdata[13][3]) # 階数（地上）
            if len(csvdata[13]) >= 6:
                excel_workbook['0) 基本情報'].cell(row=14, column=6, value=csvdata[13][5]) # 階数（地下）

            excel_workbook['0) 基本情報'].cell(row=15, column=3, value=csvdata[14][2])  # 敷地面積
            excel_workbook['0) 基本情報'].cell(row=16, column=3, value=csvdata[15][2])  # 建築面積
            excel_workbook['0) 基本情報'].cell(row=17, column=3, value=csvdata[16][2])  # 延べ面積
            excel_workbook['0) 基本情報'].cell(row=18, column=3, value=csvdata[17][2])  # 年間日射地域区分
            excel_workbook['0) 基本情報'].cell(row=19, column=3, value=csvdata[18][2])  # 一次エネ換算係数（冷熱）
            excel_workbook['0) 基本情報'].cell(row=20, column=3, value=csvdata[19][2])  # 一次エネ換算係数（温熱）
        
        elif "様式1" in csvfile_name:
            write_list_2d(excel_workbook['1) 室仕様'], csvdata[10:], 11, 1)

        elif "様式2-1" in csvfile_name or "AirConditioningRoom" in csvfile_name:
            write_list_2d(excel_workbook['2-1) 空調ゾーン'], csvdata[10:], 11, 1)

        elif "様式2-2" in csvfile_name or "WallConfiguration" in csvfile_name:
            write_list_2d(excel_workbook['2-2) 外壁構成 '], csvdata[10:], 11, 1)

        elif "様式2-3" in csvfile_name or "WindowConfiguration" in csvfile_name:
            write_list_2d(excel_workbook['2-3) 窓仕様'], csvdata[10:], 11, 1)

        elif "様式2-4" in csvfile_name or "Envelope" in csvfile_name:
            write_list_2d(excel_workbook['2-4) 外皮 '], csvdata[10:], 11, 1)

        elif "様式2-5" in csvfile_name or "HeatSource" in csvfile_name:
            write_list_2d(excel_workbook['2-5) 熱源'], csvdata[10:], 11, 1)

        elif "様式2-6" in csvfile_name or "SecondaryPump" in csvfile_name:
            write_list_2d(excel_workbook['2-6) 2次ﾎﾟﾝﾌﾟ'], csvdata[10:], 11, 1)

        elif "様式2-7" in csvfile_name or "AirHandlingUnit" in csvfile_name:
            write_list_2d(excel_workbook['2-7) 空調機'], csvdata[10:], 11, 1)

        elif "様式3-1" in csvfile_name:
            write_list_2d(excel_workbook['3-1) 換気室'], csvdata[10:], 11, 1)

        elif "様式3-2" in csvfile_name:
            write_list_2d(excel_workbook['3-2) 換気送風機'], csvdata[10:], 11, 1)

        elif "様式3-3" in csvfile_name:
            write_list_2d(excel_workbook['3-3) 換気空調機'], csvdata[10:], 11, 1)

        elif "様式4" in csvfile_name:
            write_list_2d(excel_workbook['4) 照明'], csvdata[10:], 11, 1)

        elif "様式5-1" in csvfile_name:
            write_list_2d(excel_workbook['5-1) 給湯室'], csvdata[10:], 11, 1)

        elif "様式5-2" in csvfile_name:
            write_list_2d(excel_workbook['5-2) 給湯機器'], csvdata[10:], 11, 1)

        elif "様式6" in csvfile_name:
            write_list_2d(excel_workbook['6) 昇降機'], csvdata[10:], 11, 1)

        elif "様式7-1" in csvfile_name:
            write_list_2d(excel_workbook['7-1) 太陽光発電'], csvdata[10:], 11, 1)

        elif "様式7-3" in csvfile_name:
            write_list_2d(excel_workbook['7-3) コージェネレーション設備'], csvdata[10:], 11, 1)

        elif "様式8" in csvfile_name:
            write_list_2d(excel_workbook['8) 非空調外皮'], csvdata[10:], 11, 1)

        elif "様式9" in csvfile_name:
            write_list_2d(excel_workbook['9) モデル建物'], csvdata[10:], 11, 1)

    # 出力
    excel_workbook.save(output_filename + ".xlsm")


if __name__ == '__main__':

    # テスト変換（延べ面積の入力も必要）
    csv_directory = "./tests_old/ventilation_comprehensive"
    excel_filename = "Case_vt_comprehensive_"
    for i in range(0,1):
        if i < 10:
            csv2excel(csv_directory + "/Case0"+str(i), excel_filename + "0"+str(i))
        else:
            csv2excel(csv_directory + "/Case"+str(i), excel_filename + str(i))
